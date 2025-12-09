from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import User,UserProject,Month, CalendarWeek,Task,MonitoringEntry,PlannedDedication
from django.contrib import messages
from django.db.models import Q
from .models import Project
from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Sum
import os
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.db import models



def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.role == "manager":
                return redirect("manager_dashboard")
            else:
                return redirect("employee_dashboard")
        else:
            return render(request, "login.html", {"error": "Invalid username or password"})

    return render(request, "login.html")

@login_required
def employee_dashboard(request):
    return render(request, "employee_dashboard.html")

@login_required
def manager_dashboard(request):
    # Count number of employees
    employees_count = User.objects.filter(role="employee").count()
    projects_count = Project.objects.count()

    return render(request, "manager_dashboard.html", {
        "employees_count": employees_count,
        "projects_count": projects_count
    })

def logout_view(request):
    logout(request)
    return redirect("login")

@login_required
def manage_employees(request):
    # Only managers can access this page
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    employees = User.objects.all()   # ✅ now fetches all users
    employees_count = employees.count()

    return render(request, "manage_employees.html", {
        "employees": employees,
        "employees_count": employees_count
    })


@login_required
def add_employee(request):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        is_special = request.POST.get("special_user") == "on"  # checkbox

        if username and password and email and first_name and last_name:
            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists")
            else:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                )
                user.role = "employee"
                user.special_user = 1 if is_special else 0
                user.save()
                return redirect("manage_employees")
        else:
            messages.error(request, "Please fill all fields")

    return render(request, "add_employee.html")



@login_required
def edit_employee(request, user_id):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    try:
        employee = User.objects.get(id=user_id, role="employee")
    except User.DoesNotExist:
        messages.error(request, "Employee not found")
        return redirect("manage_employees")

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")

        if username:
            employee.username = username
        if email:
            employee.email = email
        if password:
            employee.set_password(password)  # important for Django auth
        employee.save()
        ##messages.success(request, "Employee updated successfully")
        return redirect("manage_employees")

    return render(request, "edit_employee.html", {"employee": employee})

@login_required
def delete_employee(request, user_id):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    try:
        employee = User.objects.get(id=user_id, role="employee")
        employee.delete()
        ##messages.success(request, "Employee deleted successfully")
    except User.DoesNotExist:
        messages.error(request, "Employee not found")

    return redirect("manage_employees")


@login_required
def manage_projects(request):
    projects = Project.objects.all()
    return render(request, "manage_projects.html", {"projects": projects})

@login_required
def add_project(request):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    if request.method == "POST":
        name = request.POST.get("name")
        manager_id = request.POST.get("manager")  # dropdown with manager IDs

        if name and manager_id:
            manager = User.objects.filter(id=manager_id, role="manager").first()
            project = Project.objects.create(name=name, manager=manager)
            return redirect("manage_projects")
        else:
            messages.error(request, "Please fill all fields")

    managers = User.objects.filter(role="manager")
    return render(request, "add_project.html", {"managers": managers})

@login_required
def edit_project(request, project_id):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        messages.error(request, "Project not found")
        return redirect("manage_projects")

    if request.method == "POST":
        name = request.POST.get("name")
        manager_id = request.POST.get("manager")

        if name:
            project.name = name
        if manager_id:
            manager = User.objects.filter(id=manager_id, role="manager").first()
            project.manager = manager
        project.save()
        return redirect("manage_projects")

    managers = User.objects.filter(role="manager")
    return render(request, "edit_project.html", {"project": project, "managers": managers})

@login_required
def delete_project(request, project_id):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    try:
        project = Project.objects.get(id=project_id)
        project.delete()
    except Project.DoesNotExist:
        messages.error(request, "Project not found")

    return redirect("manage_projects")

@login_required
def manage_resources(request):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    # Get all users to populate the filter dropdown
    users = User.objects.all()

    # Fetch resources with employee + project in one query
    resources = UserProject.objects.select_related("employee", "project").all()

    # Apply user filter if selected
    user_id = request.GET.get("user")
    if user_id:
        resources = resources.filter(employee_id=user_id)

    context = {
        "resources": resources,
        "users": users
    }
    return render(request, "manage_resources.html", context)



@login_required
def add_assigned_resource(request):
    employees = User.objects.all()
    projects = Project.objects.all()

    if request.method == "POST":
        employee_id = request.POST.get("employee")
        project_id = request.POST.get("project")

        if employee_id and project_id:
            try:
                employee = User.objects.get(id=employee_id)
                project = Project.objects.get(id=project_id)

                # Only create with employee and project, no dedication
                if UserProject.objects.filter(employee=employee, project=project).exists():
                    messages.warning(request, "This resource is already assigned to the project.")
                else:
                    UserProject.objects.create(
                        employee=employee,
                        project=project
                    )
                    #messages.success(request, "Resource assigned successfully!")

                return redirect("manage_resources")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    return render(request, "add_assigned_resource.html", {
        "employees": employees,
        "projects": projects,
    })





from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import UserProject, User, Project

@login_required
def manage_resources(request):
    if request.user.role != "manager":
        return redirect("employee_dashboard")

    # Get all users to populate the filter dropdown
    users = User.objects.all()

    # Fetch resources with employee + project in one query
    resources = UserProject.objects.select_related("employee", "project").all()

    # Apply user filter if selected
    user_id = request.GET.get("user")
    if user_id:
        resources = resources.filter(employee_id=user_id)

    context = {
        "resources": resources,
        "users": users
    }
    return render(request, "manage_resources.html", context)

# Add new resource assignment
@login_required
def add_assigned_resource(request):
    employees = User.objects.all()
    projects = Project.objects.all()

    if request.method == "POST":
        employee_id = request.POST.get("employee")
        project_id = request.POST.get("project")

        if employee_id and project_id:
            try:
                employee = User.objects.get(id=employee_id)
                project = Project.objects.get(id=project_id)

                # Only create with employee and project, no dedication
                if UserProject.objects.filter(employee=employee, project=project).exists():
                    messages.warning(request, "This resource is already assigned to the project.")
                else:
                    UserProject.objects.create(
                        employee=employee,
                        project=project
                    )
                    #messages.success(request, "Resource assigned successfully!")

                return redirect("manage_resources")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    return render(request, "add_assigned_resource.html", {
        "employees": employees,
        "projects": projects,
    })


# Edit existing resource
@login_required
def edit_assigned_resource(request, resource_id):
    resource = get_object_or_404(UserProject, id=resource_id)
    employees = User.objects.all()
    projects = Project.objects.all()

    if request.method == "POST":
        employee_id = request.POST.get("employee")
        project_id = request.POST.get("project")

        resource.employee = get_object_or_404(User, id=employee_id)
        resource.project = get_object_or_404(Project, id=project_id)
        resource.save()

        
        return redirect("manage_resources")

    return render(request, "edit_assigned_resource.html", {
        "resource": resource,
        "employees": employees,
        "projects": projects
    })

# Delete resource
def delete_assigned_resource(request, resource_id):
    resource = get_object_or_404(UserProject, id=resource_id)
    resource.delete()
    ##messages.success(request, "Resource deleted successfully.")
    return redirect("manage_resources")

def manage_cw(request):
    MONTH_ORDER = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    months = list(Month.objects.all())
    months.sort(key=lambda m: MONTH_ORDER.index(m.month))
    calendar_weeks = CalendarWeek.objects.all()

    if request.method == "POST":
        month_id = request.POST.get("month_id")
        cw_ids = request.POST.getlist("cw_ids")  # multiple values

        if month_id and cw_ids:
            for cw_id in cw_ids:
                cw = get_object_or_404(CalendarWeek, id=cw_id)
                cw.month_id = month_id
                cw.save()
            return redirect("manage_cw")

    return render(request, "manage_cw.html", {"months": months, "calendar_weeks": calendar_weeks})


def manage_tasks(request):
    tasks = Task.objects.select_related("project").all()  # optimize query
    return render(request, "manage_tasks.html", {"tasks": tasks})

def add_task(request):
    projects = Project.objects.all()
    if request.method == "POST":
        task_name = request.POST.get("task_name")
        project_id = request.POST.get("project")
        common_task = request.POST.get("common_task") == "on"

        if common_task:
            Task.objects.create(task_name=task_name, common_task=True)
        else:
            project = Project.objects.get(id=project_id)
            Task.objects.create(task_name=task_name, project=project)

        return redirect("manage_tasks")

    return render(request, "add_task.html", {"projects": projects})

def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    projects = Project.objects.all()
    
    if request.method == "POST":
        task_name = request.POST.get("task_name")
        project_id = request.POST.get("project")
        if project_id:
            project = Project.objects.get(id=project_id)
        else:
            project = None  # task common to all projects
        
        task.task_name = task_name
        task.project = project
        task.save()
        ##messages.success(request, "Task updated successfully!")
        return redirect('manage_tasks')
    
    return render(request, "edit_task.html", {"task": task, "projects": projects})

# Delete Task
def delete_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.delete()
    ##messages.success(request, "Task deleted successfully!")
    return redirect('manage_tasks')

@login_required
def fill_monitoring(request):
    months = Month.objects.all()
    cws = CalendarWeek.objects.select_related('month').all()

    # Debug: which projects does this user have?
    user_projects = UserProject.objects.filter(employee=request.user).select_related("project")
    project_ids = [up.project.id for up in user_projects]

    print("User:", request.user.username)
    print("User projects:", project_ids)

    # Fetch tasks
    tasks = Task.objects.filter(Q(project_id__in=project_ids) | Q(common_task=True)).select_related("project")

    print("Tasks count:", tasks.count())
    for t in tasks:
        print("Task:", t.task_name, "Project:", t.project)

    if request.method == 'POST':
        month_id = request.POST.get('month')
        cw_id = request.POST.get('cw')

        for task in tasks:
            hours = request.POST.get(f'hours_{task.id}')
            if hours and month_id and cw_id:
                MonitoringEntry.objects.update_or_create(
                    user=request.user,
                    task=task,
                    cw_id=int(cw_id),
                    defaults={
                        'project': task.project if task.project else None,
                        'month_id': int(month_id),
                        'hours_spent': float(hours)
                    }
                )
        return redirect('monitoring_dashboard')

    return render(request, 'fill_monitoring.html', {
        'months': months,
        'cws': cws,
        'tasks': tasks,
    })



@login_required
def monitoring_success(request):
    return render(request, 'monitoring_success.html')


@login_required


@login_required
def monitoring_dashboard(request):
    all_cws = CalendarWeek.objects.select_related("month").all().order_by("month__id", "cw")

    monitoring_entries = (
        MonitoringEntry.objects
        .filter(user=request.user)
        .select_related("month", "cw", "project", "task")
    )

    # Filter by CW if selected
    cw_id = request.GET.get("cw")
    if cw_id:
        monitoring_entries = monitoring_entries.filter(cw_id=cw_id)

    # ---- 🔽 Sort by the number after "CW" ----
    monitoring_entries = sorted(
        monitoring_entries,
        key=lambda e: int(e.cw.cw.replace("CW", "")) if e.cw and e.cw.cw else 0
    )
    # -------------------------------------------

    return render(
        request,
        "monitoring_dashboard.html",
        {
            "monitoring_entries": monitoring_entries,
            "all_cws": all_cws,
        },
    )


@login_required
def consult_monitoring(request):
    # Restrict access to managers
    if request.user.role != "manager":
        return redirect("monitoring_dashboard")

    users = User.objects.all().order_by("username")
    cws = CalendarWeek.objects.select_related("month").all()
    entries = []
    selected_user = None
    selected_cw = None
    total_hours = 0

    if request.method == "POST":
        user_id = request.POST.get("user")
        cw_id = request.POST.get("cw")

        if user_id and cw_id:
            selected_user = User.objects.get(id=user_id)
            selected_cw = CalendarWeek.objects.get(id=cw_id)

            # ✅ 1. Get all monitoring entries filled by this user in that CW
            filled_entries = (
                MonitoringEntry.objects
                .filter(user=selected_user, cw=selected_cw)
                .select_related("task", "project", "month")
            )

            # ✅ 2. Get all projects linked to the user
            linked_projects = Project.objects.filter(user_projects__employee=selected_user)

            # ✅ 3. Get all tasks from those projects + all common tasks
            all_tasks = Task.objects.filter(
                models.Q(project__in=linked_projects) | models.Q(common_task=True)
            ).select_related("project")

            # ✅ 4. Map filled tasks by ID for quick lookup
            filled_map = {entry.task.id: entry for entry in filled_entries}

            # ✅ 5. Combine filled + not-filled (linked + common tasks)
            entries = []
            total_hours = 0

            for task in all_tasks:
                if task.id in filled_map:
                    entry = filled_map[task.id]
                    entries.append(entry)
                    total_hours += entry.hours_spent
                else:
                    # Not filled → show as 0 hours
                    entries.append({
                        "task": task,
                        "project": task.project,
                        "month": selected_cw.month,
                        "hours_spent": 0,
                        "is_unfilled": True  # flag for styling in template
                    })

    return render(
        request,
        "consult_monitoring.html",
        {
            "users": users,
            "cws": cws,
            "entries": entries,
            "selected_user": selected_user,
            "selected_cw": selected_cw,
            "total_hours": total_hours,
        },
    )

from collections import defaultdict
@login_required
def export_monitoring_excel(request):
    # Define month order
    MONTH_ORDER = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    # Get all months and sort
    months = list(Month.objects.all())
    months.sort(key=lambda m: MONTH_ORDER.index(m.month))
    
    month_id = request.GET.get("month")

    if month_id:
        try:
            month = Month.objects.get(id=month_id)

            # 1️⃣ Exclude entries that belong to common tasks
            entries = (
                MonitoringEntry.objects
                .filter(month=month)
                .exclude(task__common_task=True)
                .select_related("project", "user", "task")
            )

            # 2️⃣ Count number of CWs in the selected month
            num_cws = CalendarWeek.objects.filter(month=month).count()

            # 3️⃣ Total worked hours per user (excluding common tasks)
            user_totals_qs = (
                entries.values("user__id", "user__first_name", "user__last_name", "user__username", "user__special_user")
                .annotate(total_hours=Sum("hours_spent"))
            )
            totals_map = {
                row["user__id"]: {
                    "username": f"{row['user__first_name']} {row['user__last_name']}".strip() or row.get('user__username', ''),
                    "total_hours": float(row["total_hours"] or 0),
                    "special_user": row["user__special_user"]
                }
                for row in user_totals_qs
            }

            # 4️⃣ Aggregate hours per project/user
            project_user_data = (
                entries.values("project__name", "user__id")
                .annotate(project_hours=Sum("hours_spent"))
                .order_by("project__name", "user__id")
            )

            # 5️⃣ Group by project
            grouped = defaultdict(list)
            for row in project_user_data:
                grouped[row["project__name"]].append(row)

            # 6️⃣ Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Monitoring {month.month}"

            # 🔹 UPDATED HEADER: Month BEFORE Hours
            headers = ["Project", "Full Name", "Month", "Hours", "%"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            current_row = 2
            for project, rows in grouped.items():
                start_row = current_row
                for row in rows:
                    user_info = totals_map[row["user__id"]]
                    username = user_info["username"]
                    project_hours = float(row["project_hours"] or 0)
                    total_hours_user = user_info["total_hours"]
                    special_user = user_info["special_user"]

                    # Base hours for percentage
                    if special_user:
                        if num_cws == 4:
                            base_hours = 160
                        elif num_cws == 5:
                            base_hours = 200
                        else:
                            base_hours = total_hours_user
                    else:
                        base_hours = total_hours_user

                    percentage = 0
                    if base_hours > 0:
                        percentage = round(((project_hours / base_hours) * 100) / 5) * 5

                    # 🔹 UPDATED ROW FORMAT: Month BEFORE Hours
                    ws.append([project, username, month.month, project_hours, f"{percentage}%"])
                    current_row += 1

                # Merge project cells
                if len(rows) > 1:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=1,
                        end_row=current_row - 1,
                        end_column=1
                    )

            # 7️⃣ Return Excel file
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"Monitoring_{month.month}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response

        except Month.DoesNotExist:
            pass

    return render(request, "export_monitoring.html", {"months": months})






@login_required
def user_monitoring_dashboard(request):
    cw_filter = request.GET.get('cw')
    entries = MonitoringEntry.objects.filter(user=request.user).select_related('month', 'cw', 'project', 'task')
    
    if cw_filter:
        entries = entries.filter(cw_id=cw_filter)
    
    # Sort entries by CW number
    entries = sorted(
        entries,
        key=lambda e: int(e.cw.cw.replace('CW', '')) if e.cw.cw.startswith('CW') else 0
    )
    
    all_cws = CalendarWeek.objects.all()
    
    return render(request, 'user_monitoring_dashboard.html', {
        'monitoring_entries': entries,
        'all_cws': all_cws,
    })

@login_required
def user_fill_monitoring(request):
    months = Month.objects.all()
    cws = CalendarWeek.objects.select_related('month').all()

    # User projects
    user_projects = UserProject.objects.filter(employee=request.user).select_related("project")
    project_ids = [up.project.id for up in user_projects]

    # Tasks related to user projects OR common tasks
    tasks = Task.objects.filter(Q(project_id__in=project_ids) | Q(common_task=True)).select_related("project")

    if request.method == 'POST':
        month_id = request.POST.get('month')
        cw_id = request.POST.get('cw')

        for task in tasks:
            hours = request.POST.get(f'hours_{task.id}')
            if hours and month_id and cw_id:
                MonitoringEntry.objects.update_or_create(
                    user=request.user,
                    task=task,
                    cw_id=int(cw_id),
                    defaults={
                        'project': task.project if task.project else None,
                        'month_id': int(month_id),
                        'hours_spent': float(hours)
                    }
                )
        return redirect('user_monitoring_dashboard')

    return render(request, 'user_fill_monitoring.html', {
        'months': months,
        'cws': cws,
        'tasks': tasks,
    })

@login_required
def user_edit_profile(request):
    user = request.user

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')

        # Update username/email
        user.username = username
        user.email = email
        user.save()

        # Update password if provided
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
            user.save()
            # Keep user logged in after password change
            update_session_auth_hash(request, user)

        ##messages.success(request, "Profile updated successfully!")
        return redirect('user_edit_profile')

    return render(request, 'user_edit_profile.html', {'user': user})

@login_required
def edit_monitoring(request, entry_id):
    entry = get_object_or_404(MonitoringEntry, id=entry_id, user=request.user)

    if request.method == "POST":
        project_id = request.POST.get("project")
        task_id = request.POST.get("task")
        hours_spent = request.POST.get("hours_spent")

        entry.project_id = project_id if project_id else None
        entry.task_id = task_id
        entry.hours_spent = hours_spent
        entry.save()

        ##messages.success(request, "✅ Monitoring entry updated successfully.")
        return redirect("monitoring_dashboard")

    projects = Project.objects.all()
    tasks = Task.objects.all()
    return render(request, "edit_monitoring.html", {
        "entry": entry,
        "projects": projects,
        "tasks": tasks,
    })


@login_required
def delete_monitoring(request, entry_id):
    entry = get_object_or_404(MonitoringEntry, id=entry_id, user=request.user)
    entry.delete()
    ##messages.success(request, "🗑️ Monitoring entry deleted successfully.")
    return redirect("monitoring_dashboard")

def user_edit_monitoring(request, entry_id):
    entry = get_object_or_404(MonitoringEntry, id=entry_id, user=request.user)

    if request.method == "POST":
        project_id = request.POST.get("project")
        task_id = request.POST.get("task")
        month_id = request.POST.get("month")
        cw_id = request.POST.get("cw")
        hours_spent = request.POST.get("hours_spent")

        # Update fields
        entry.project_id = project_id if project_id else None
        entry.task_id = task_id
        entry.month_id = month_id
        entry.cw_id = cw_id
        entry.hours_spent = hours_spent
        entry.save()

        ##messages.success(request, "Monitoring entry updated successfully.")
        return redirect("user_monitoring_dashboard")

    context = {
        "entry": entry,
        "projects": Project.objects.all(),
        "tasks": Task.objects.all(),
        "months": Month.objects.all(),
        "cws": CalendarWeek.objects.all(),
    }
    return render(request, "user_edit_monitoring.html", context)


@login_required
def user_delete_monitoring(request, entry_id):
    entry = get_object_or_404(MonitoringEntry, id=entry_id, user=request.user)
    entry.delete()
    ##messages.success(request, "Monitoring entry deleted successfully.")
    return redirect("user_monitoring_dashboard")

@login_required
def manager_edit_profile(request):
    user = request.user  # manager user

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")

        # Update username and email
        user.username = username
        user.email = email

        # Update password only if provided
        if password:
            user.set_password(password)

        user.save()

        #messages.success(request, "Profile updated successfully!")

        # If password changed, log the user in again
        if password:
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)

        return redirect('manager_edit_profile')

    return render(request, "manager_edit_profile.html", {"user": user})

@login_required
def consult_monitoring_cw(request):
    """
    Manager page:
    - Select a CW from a dropdown
    - Show all MonitoringEntry rows for that CW
    - Merge 'User' column
    - Add total hours per user
    """
    # Define the fixed CW order
    CW_ORDER = [f"CW{i}" for i in range(1, 53)]  # CW1 → CW52

    # Retrieve all CWs and manually sort using CW_ORDER
    cws = list(CalendarWeek.objects.all())
    cws.sort(key=lambda cw: CW_ORDER.index(cw.cw) if cw.cw in CW_ORDER else 999)

    selected_cw = request.GET.get("cw")

    entries_qs = MonitoringEntry.objects.none()
    rows = []

    if selected_cw:
        try:
            cw_id = int(selected_cw)
            entries_qs = (
                MonitoringEntry.objects
                .filter(cw_id=cw_id)
                .select_related("user", "task__project")
                .order_by("user__username", "task__task_name")
            )

            if entries_qs.exists():
                buffer = []
                current_user = None
                count = 0
                user_total_hours = 0

                for entry in entries_qs:
                    if current_user != entry.user.username:
                        if buffer:
                            for i, e in enumerate(buffer):
                                rows.append({
                                    "entry": e,
                                    "rowspan": count if i == 0 else 0,
                                    "total_hours": user_total_hours if i == 0 else 0
                                })
                        buffer = [entry]
                        current_user = entry.user.username
                        count = 1
                        user_total_hours = entry.hours_spent
                    else:
                        buffer.append(entry)
                        count += 1
                        user_total_hours += entry.hours_spent

                if buffer:
                    for i, e in enumerate(buffer):
                        rows.append({
                            "entry": e,
                            "rowspan": count if i == 0 else 0,
                            "total_hours": user_total_hours if i == 0 else 0
                        })

        except (ValueError, TypeError):
            pass

    return render(
        request,
        "consult_monitoring_cw.html",
        {
            "cws": cws,
            "selected_cw": selected_cw,
            "rows": rows,
            "has_entries": entries_qs.exists(),
        },
    )

from collections import defaultdict, OrderedDict



@login_required
def planned_dedication_list(request):
    dedications = PlannedDedication.objects.all().select_related("user", "project", "month")

    # Apply filters
    month_id = request.GET.get("month")
    user_id = request.GET.get("user")
    project_id = request.GET.get("project")
    dedication_val = request.GET.get("dedication")

    if month_id:
        dedications = dedications.filter(month_id=month_id)
    if user_id:
        dedications = dedications.filter(user_id=user_id)
    if project_id:
        dedications = dedications.filter(project_id=project_id)
    if dedication_val:
        dedications = dedications.filter(planned_dedication=dedication_val)

    # Manual month ordering
    MONTH_ORDER = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    months = list(Month.objects.all())
    months.sort(key=lambda m: MONTH_ORDER.index(m.month))

    # 🧩 Sort dedications by user username and month order
    dedications = sorted(
        dedications,
        key=lambda d: (d.user.username.lower(), MONTH_ORDER.index(d.month.month))
    )

    context = {
        "dedications": dedications,
        "months": months,
        "users": User.objects.all(),
        "projects": Project.objects.all(),
    }
    return render(request, "planned_dedication_list.html", context)


@login_required
def add_planned_dedication(request):
    MONTH_ORDER = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    users = User.objects.all()
    months = list(Month.objects.all())
    months.sort(key=lambda m: MONTH_ORDER.index(m.month))

    if request.method == "POST":
        user_id = request.POST.get('user')
        if not user_id:
            messages.error(request, "Please select a user.")
            return redirect("add_planned_dedication")

        user = get_object_or_404(User, id=user_id)

        for key, value in request.POST.items():
            if key.startswith("dedication_"):  # e.g., dedication_0_1
                try:
                    dedication_val = int(value)
                except ValueError:
                    messages.error(request, "Dedication must be a number.")
                    return redirect("add_planned_dedication")

                # Extract project index and month id from key
                _, project_index, month_id = key.split("_")
                month = get_object_or_404(Month, id=month_id)

                # Get project id from hidden input in the same index
                project_key = f"project_{project_index}"
                project_id = request.POST.get(project_key)
                if not project_id:
                    messages.error(request, "Project not found for dedication input.")
                    return redirect("add_planned_dedication")
                project = get_object_or_404(Project, id=project_id)

                # Calculate existing dedication for this user and month
                existing_total = PlannedDedication.objects.filter(
                    user=user, month=month
                ).aggregate(total=models.Sum("planned_dedication"))["total"] or 0

                if existing_total + dedication_val > 100:
                    messages.error(
                        request,
                        f"Total dedication for {month.month} cannot exceed 100% (currently: {existing_total}%)."
                    )
                    return redirect("add_planned_dedication")

                # Save the dedication with project
                PlannedDedication.objects.create(
                    user=user,
                    project=project,
                    month=month,
                    planned_dedication=dedication_val
                )

        messages.success(request, f"Planned dedication added for {user.username}.")
        return redirect("planned_dedication_list")

    # GET request
    context = {
        "users": users,
        "months": months,
    }
    return render(request, "add_planned_dedication.html", context)

# AJAX endpoint: get projects assigned to a user
@login_required
def get_user_projects(request):
    user_id = request.GET.get("user_id")
    projects = []
    if user_id:
        user_projects = UserProject.objects.filter(employee_id=user_id).select_related("project")
        for up in user_projects:
            projects.append({"id": up.project.id, "name": up.project.name})
    return JsonResponse({"projects": projects})



from openpyxl.utils import get_column_letter
from openpyxl import Workbook
@login_required
def export_monthly_dedication(request):
    months = Month.objects.all()
    MONTH_ORDER = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    # Get all months and sort
    months = list(Month.objects.all())
    months.sort(key=lambda m: MONTH_ORDER.index(m.month))

    if request.method == "POST":
        month_id = request.POST.get("month")
        try:
            month_obj = Month.objects.get(id=month_id)
        except Month.DoesNotExist:
            return render(request, "export_monthly_dedication.html", {"months": months, "error": "Selected month does not exist."})

        # 1️⃣ Fetch planned dedication
        planned_qs = PlannedDedication.objects.filter(month=month_obj).select_related("user", "project")

        # 2️⃣ Fetch real dedication from MonitoringEntry
        entries = (
            MonitoringEntry.objects
            .filter(month=month_obj)
            .exclude(task__common_task=True)
            .select_related("user", "project", "task")
        )

        num_cws = CalendarWeek.objects.filter(month=month_obj).count()

        # Total hours per user
        totals_qs = entries.values(
            "user__id", "user__first_name", "user__last_name",
            "user__username", "user__special_user"
        ).annotate(total_hours=Sum("hours_spent"))

        totals_map = {
            row["user__id"]: {
                "username": f"{row['user__first_name']} {row['user__last_name']}".strip() or row["user__username"],
                "total_hours": float(row["total_hours"] or 0),
                "special_user": row["user__special_user"]
            }
            for row in totals_qs
        }

        # Map hours per project/user
        hours_qs = entries.values("project__id", "user__id", "project__name").annotate(project_hours=Sum("hours_spent"))
        hours_map = {(row["project__id"], row["user__id"]): float(row["project_hours"] or 0) for row in hours_qs}

        # Group users under each project
        projects_users_map = defaultdict(list)
        all_keys = set(list(hours_map.keys()) + [(d.project.id, d.user.id) for d in planned_qs])

        for project_id, user_id in all_keys:
            projects_users_map[project_id].append(user_id)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Dedications {month_obj.month}"

        # 🔹 HEADER (unchanged)
        headers = ["Project", "Full Name", "Month", "Planned Dedication", "Real Dedication"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        current_row = 2

        for project_id, user_ids in projects_users_map.items():
            start_row = current_row

            # Get project name
            project_name_obj = planned_qs.filter(project_id=project_id).first()
            if project_name_obj:
                project_name = project_name_obj.project.name
            else:
                any_hours_entry = next(((pid, uid) for (pid, uid) in hours_map if pid == project_id), None)
                if any_hours_entry:
                    project_name = MonitoringEntry.objects.filter(project_id=project_id).first().project.name
                else:
                    project_name = "Unknown Project"

            for user_id in user_ids:

                # Planned
                planned_entry = planned_qs.filter(user_id=user_id, project_id=project_id).first()
                planned_value = float(planned_entry.planned_dedication or 0) if planned_entry else 0

                # Real
                project_hours = hours_map.get((project_id, user_id), 0)

                if user_id in totals_map:
                    user_info = totals_map[user_id]
                    username = user_info["username"]
                    total_hours = user_info["total_hours"]
                    special_user = user_info["special_user"]
                elif planned_entry:
                    username = f"{planned_entry.user.first_name} {planned_entry.user.last_name}".strip() or planned_entry.user.username
                    total_hours = project_hours
                    special_user = False
                else:
                    username = "Unknown"
                    total_hours = project_hours
                    special_user = False

                # % Calculation
                if special_user:
                    if num_cws == 4:
                        base_hours = 160
                    elif num_cws == 5:
                        base_hours = 200
                    else:
                        base_hours = total_hours
                else:
                    base_hours = total_hours

                if base_hours > 0:
                    real_pct = round(((project_hours / base_hours) * 100) / 5) * 5
                else:
                    real_pct = 0

                # 🔹 BOTH planned and real dedication now include %  
                ws.append([
                    project_name,
                    username,
                    month_obj.month,
                    f"{planned_value}%",     # ← ADDED %
                    f"{real_pct}%",          # Already had %
                ])

                current_row += 1

            # Merge project name cells
            if len(user_ids) > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)

        # Column widths
        for i, col in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        # Return file
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"Monthly_Dedication_{month_obj.month}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    return render(request, "export_monthly_dedication.html", {"months": months})



@login_required
def edit_planned_dedication(request, pk):
    planned = get_object_or_404(PlannedDedication, id=pk)

    if request.method == "POST":
        user_id = request.POST.get("user")
        project_id = request.POST.get("project")
        month_id = request.POST.get("month")
        planned_dedication_val = request.POST.get("planned_dedication")

        # Validate and update
        user = get_object_or_404(User, id=user_id)
        project = get_object_or_404(Project, id=project_id)
        month = get_object_or_404(Month, id=month_id)

        planned.user = user
        planned.project = project
        planned.month = month
        planned.planned_dedication = planned_dedication_val
        planned.save()

        return redirect("planned_dedication_list")

    context = {
        "planned": planned,
        "users": User.objects.all(),
        "projects": Project.objects.all(),
        "months": Month.objects.all(),
    }
    return render(request, "edit_planned_dedication.html", context)

@login_required
def delete_planned_dedication(request, pk):
    planned = get_object_or_404(PlannedDedication, id=pk)
    planned.delete()
    return redirect("planned_dedication_list")


from django.db.models import Sum
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
@login_required

def export_overview_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Planned Dedication Overview"

    # Header row
    months = Month.objects.order_by("id")
    header = ["User"] + [m.month for m in months]
    ws.append(header)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # red background

    # Style header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows: all employees
    users = User.objects.filter(role="employee").order_by("username")
    for user in users:
        row = [user.username]
        for month in months:
            total = PlannedDedication.objects.filter(user=user, month=month).aggregate(
                total=Sum("planned_dedication")
            )["total"] or 0
            row.append(total)
        ws.append(row)

    # Apply styles to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        username_cell = row[0]
        username_cell.font = Font(bold=True)
        username_cell.alignment = Alignment(horizontal="left")
        username_cell.border = thin_border

        for cell in row[1:]:
            cell.alignment = center_align
            cell.border = thin_border
            # Highlight in red if total is not 100
            if isinstance(cell.value, (int, float)) and cell.value != 100:
                cell.fill = red_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 4

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="planned_dedication_overview.xlsx"'
    wb.save(response)
    return response

@login_required
def export_resources(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["First Name", "Last Name"]
    ws.append(headers)

    # Apply style to header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Fetch users and add rows
    users = User.objects.all().values_list("first_name", "last_name")
    for user in users:
        ws.append(user)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    # Prepare response for download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="resources.xlsx"'

    wb.save(response)
    return response

def export_projects_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # Add header
    ws.append(["Project"])

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    header_cell = ws["A1"]
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = center_align
    header_cell.border = thin_border

    # Add all projects
    projects = Project.objects.all().order_by("name")
    for project in projects:
        ws.append([project.name])

    # Style project names
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Auto width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 5

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="projects_list.xlsx"'
    wb.save(response)
    return response



MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

@login_required
def employee_consult_planned_dedication(request):
    user = request.user

    # Get user's assigned projects
    user_projects = [up.project for up in user.user_projects.all()]

    # All months
    months = Month.objects.all()

    # Filters
    selected_project = request.GET.get("project", "all")
    selected_month = request.GET.get("month", "all")

    dedications = PlannedDedication.objects.filter(user=user)

    if selected_project != "all":
        dedications = dedications.filter(project_id=selected_project)
    if selected_month != "all":
        dedications = dedications.filter(month_id=selected_month)

    # Order months using MONTH_ORDER
    dedications = sorted(
        dedications,
        key=lambda d: MONTH_ORDER.index(d.month.month) if d.month.month in MONTH_ORDER else 12
    )

    context = {
        "user_projects": user_projects,
        "months": months,
        "month_order": MONTH_ORDER,
        "dedications": dedications,
        "selected_project": selected_project,
        "selected_month": selected_month,
    }
    return render(request, "employee_consult_planned_dedication.html", context)


def clear_all_planned_dedications(request):
    try:
        PlannedDedication.objects.all().delete()
        #messages.success(request, "All planned dedications have been deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting records: {str(e)}")

    return redirect("planned_dedication_list")